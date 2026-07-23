"""
VDI Survey Processor – Household Survey Analytics (v3)
Reads HOUSEHOLD SURVEY + GRAM PANCHAYAT sheets.
Maps raw survey responses (Y/N, categorical, numeric) to meaningful analytics.
Never plots raw Excel cell values as chart labels.
"""
import json, re, sys
from collections import Counter
from openpyxl import load_workbook

# ──────────────────────────────────────────────────────────────────────────────
# Core helpers
# ──────────────────────────────────────────────────────────────────────────────

def nrm(v):
    """Lowercase alphanumeric-only normalisation for matching."""
    return re.sub(r"[^a-z0-9]", "", str(v or "").lower())

def worksheet(wb, name):
    n = nrm(name)
    s = next((s for s in wb.sheetnames if nrm(s) == n), None)
    if not s:
        s = next((s for s in wb.sheetnames if n in nrm(s)), None)
    if not s:
        raise ValueError(f"Sheet '{name}' not found.")
    return wb[s]

def table_rows(ws):
    """Return list[dict] from sheet, resolving merged cells, skipping metadata rows."""
    raw = list(ws.iter_rows(values_only=True))

    # Resolve merged cells
    merged = {}
    for rng in ws.merged_cells.ranges:
        top = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = top

    values = []
    for ri, row in enumerate(raw):
        values.append([merged.get((ri + 1, ci + 1), v) for ci, v in enumerate(row)])

    # Find header row: prefer ≥4 distinct non-null string values.
    header_idx = None
    header_keywords = {"s.no", "sno", "slno", "serial", "no", "id", "question",
                       "questionnaire", "indicator", "parameter", "response",
                       "answer", "value", "field", "item", "attribute"}
    for i, row in enumerate(values):
        non_null = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(set(v.lower() for v in non_null)) >= 4:
            # Verify at least one cell looks like a header keyword (skip metadata rows like name/date)
            combined = " ".join(v.lower() for v in non_null)
            if any(kw in combined for kw in header_keywords):
                header_idx = i
                break

    # Fallback for simple two-column summary sheets (e.g. Gram Panchayat Level).
    if header_idx is None and values:
        for i, row in enumerate(values):
            non_null = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if len(row) >= 2 and len(set(v.lower() for v in non_null)) >= 2:
                # Verify it looks like a table header (at least one cell has a header-like keyword)
                combined = " ".join(v.lower() for v in non_null)
                if any(kw in combined for kw in header_keywords):
                    header_idx = i
                    break

    if header_idx is None:
        return []

    headers = [str(v).strip() if v is not None else f"col_{ci}"
               for ci, v in enumerate(values[header_idx])]
    return [
        {headers[ci]: row[ci] if ci < len(row) else None
         for ci in range(len(headers))}
        for row in values[header_idx + 1:]
        if any(v is not None and str(v).strip() for v in row)
    ]

def transposed_households(ws):
    """Convert the CBRI questionnaire layout (questions in rows, households in columns) into records."""
    indicator_col = None
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if nrm(cell.value) in ("indicator", "indicators"):
                indicator_col = cell.column
                header_row = cell.row
                break
        if indicator_col:
            break
    if not indicator_col or not header_row:
        return []

    # The first answer column follows the questionnaire metadata columns.
    answer_start = indicator_col + 2
    records = []
    for column in range(answer_start, ws.max_column + 1):
        record = {}
        for row in range(header_row + 1, ws.max_row + 1):
            question = ws.cell(row, indicator_col).value
            answer = ws.cell(row, column).value
            if question not in (None, "") and answer not in (None, ""):
                record[str(question).strip()] = answer
        if record:
            records.append(record)
    return records


# ──────────────────────────────────────────────────────────────────────────────
# Metadata column detection – these columns must be skipped entirely
# ──────────────────────────────────────────────────────────────────────────────

METADATA_SKIP = {
    "serial", "sno", "srno", "slno", "no", "id", "hhid", "householdid",
    "name", "headname", "fathername", "husbandname", "guardianname", "respondent",
    "latitude", "longitude", "lat", "lon", "lng", "gps", "coordinates", "location",
    "date", "time", "timestamp", "surveyor", "enumerator", "investigator",
    "address", "village", "district", "block", "state", "tehsil", "mandal",
    "pincode", "mobile", "phone", "contact",
    "remarks", "notes", "comment", "signature", "photo",
}

def is_metadata_col(header):
    h = nrm(header)
    if h in METADATA_SKIP:
        return True
    # skip if header is purely a number or single char
    if re.fullmatch(r"\d+", h) or len(h) <= 1:
        return True
    return False


# ──────────────────────────────────────────────────────────────────────────────
# Column finder – keyword-based, skips metadata
# ──────────────────────────────────────────────────────────────────────────────

def find_col(rows, *keywords):
    """Return non-null values from the first column whose header matches any keyword."""
    if not rows:
        return []
    headers = [h for h in rows[0].keys() if not is_metadata_col(h)]
    for kw in keywords:
        kw_n = nrm(kw)
        # exact
        match = next((h for h in headers if nrm(h) == kw_n), None)
        if not match:
            # header contains keyword
            match = next((h for h in headers if kw_n in nrm(h)), None)
        if not match and len(kw_n) > 3:
            # keyword contains header
            match = next((h for h in headers if nrm(h) in kw_n), None)
        if match:
            return [r[match] for r in rows if r.get(match) is not None
                    and str(r[match]).strip() != ""]
    return []


# ──────────────────────────────────────────────────────────────────────────────
# Yes / No detection
# ──────────────────────────────────────────────────────────────────────────────

YES_VALS = {"yes", "y", "1", "1.0", "true", "ha", "haan", "हां", "हाँ", "✓", "available", "av"}
NO_VALS  = {"no", "n", "0", "0.0", "false", "nahi", "नहीं", "na", "not available", "✗"}

def is_yes(v): return str(v).strip().lower().rstrip(".") in YES_VALS
def is_no(v):  return str(v).strip().lower().rstrip(".") in NO_VALS
def is_yn(v):  return is_yes(v) or is_no(v)


# ──────────────────────────────────────────────────────────────────────────────
# Value mapping functions
# ──────────────────────────────────────────────────────────────────────────────

def map_gender(v):
    vl = str(v).strip().lower()
    if vl in ("male", "m", "पुरुष"):    return "Male"
    if vl in ("female", "f", "महिला"):  return "Female"
    #
    #if vl in ("other", "o", "others"):  return "Other"
    return None  # skip unrecognised

def map_house(v):
    vl = str(v).strip().lower()
    if "semi" in vl:                          return "Semi-Pucca"
    if any(w in vl for w in ("pucca","pakka","rcc","concrete","brick")): return "Pucca"
    if any(w in vl for w in ("kutcha","kaccha","mud","bamboo","thatch","grass","tin")): return "Kutcha"
    return None

def map_transport(v):
    vl = str(v).strip().lower()
    if any(w in vl for w in ("walk","foot","paidal","on foot")):   return "Walking"
    if any(w in vl for w in ("cycle","bicycle")):                  return "Bicycle"
    if any(w in vl for w in ("bike","motorcycle","scooter","two wheel","2-wheel","motorbike")): return "Two-Wheeler"
    if any(w in vl for w in ("car","jeep","four wheel","4-wheel","suv")): return "Car/Jeep"
    if any(w in vl for w in ("bus","auto","public","shared","tempo")): return "Public Transport"
    return None

def map_water(v):
    vl = str(v).strip().lower()
    if any(w in vl for w in ("tap","piped","nal","nali","pipeline")): return "Tap Water"
    if any(w in vl for w in ("borewell","bore","tubewell","tube well")): return "Borewell"
    if any(w in vl for w in ("hand pump","handpump","pump")):       return "Hand Pump"
    if any(w in vl for w in ("well","kuan","open well")):           return "Open Well"
    if any(w in vl for w in ("river","pond","lake","stream","nadi","talab")): return "River/Pond"
    return None

def map_fuel(v):
    vl = str(v).strip().lower()
    if any(w in vl for w in ("lpg","gas","cylinder","bottled")):     return "LPG"
    if "biogas" in vl:                                               return "Biogas"
    if any(w in vl for w in ("electric","induction")):               return "Electricity"
    if any(w in vl for w in ("wood","firewood","lkadi","biomass","agri waste","agricultural")): return "Firewood"
    if "kerosene" in vl:                                             return "Kerosene"
    if any(w in vl for w in ("dung","gobar","cow dung")):            return "Cow Dung"
    return None

def map_education(v):
    vl = str(v).strip().lower()
    if any(w in vl for w in ("illiterate","nirakshar","can't read","cannot read","0")): return "Illiterate"
    if any(w in vl for w in ("primary","class 1","class 2","class 3","class 4","class 5","1st","2nd","3rd","4th","5th")): return "Primary (I-V)"
    if any(w in vl for w in ("middle","class 6","class 7","class 8","6th","7th","8th")): return "Middle (VI-VIII)"
    if any(w in vl for w in ("secondary","class 9","class 10","matric","10th","ssc","high school")): return "Secondary (X)"
    if any(w in vl for w in ("higher secondary","class 11","class 12","12th","inter","hsc","senior secondary")): return "Senior Secondary (XII)"
    if any(w in vl for w in ("graduate","degree","ba","bsc","bcom","btech","graduation")): return "Graduate"
    if any(w in vl for w in ("post graduate","pg","master","msc","mba","phd")): return "Post-Graduate"
    if any(w in vl for w in ("literate","read","write","akshar")):  return "Literate (others)"
    return None

def map_occupation(v):
    vl = str(v).strip().lower()
    if any(w in vl for w in ("agri","farm","kisan","krishi","cultivat")):
        if any(w in vl for w in ("other","also","plus","+"," and ","labour","laborer","business","service")):
            return "Agriculture + Other"
        return "Agriculture Only"
    if any(w in vl for w in ("labour","labor","mazdoor","daily wage","casual")):
        return "Labour"
    if any(w in vl for w in ("business","shop","trade","vyapar","self employ","entrepreneur")):
        return "Business"
    if any(w in vl for w in ("service","govt","sarkari","job","employ","naukri","salary")):
        return "Service"
    if any(w in vl for w in ("pension","retired","retd","ex-service")):
        return "Retired/Pensioner"
    if any(w in vl for w in ("student","school","college")):
        return "Student"
    if any(w in vl for w in ("house","homemaker","housewife","grahini","domestic")):
        return "Homemaker"
    return "Other"

def map_livestock(v):
    vl = str(v).strip().lower()
    if any(w in vl for w in ("cow","gaay","cattle","gai")):     return "Cow"
    if any(w in vl for w in ("buffalo","bhains","bull","ox")):  return "Buffalo"
    if any(w in vl for w in ("goat","sheep","bakri")):          return "Goat/Sheep"
    if any(w in vl for w in ("poultry","hen","chicken","murgi","duck")): return "Poultry"
    if any(w in vl for w in ("pig","swine")):                   return "Pig"
    if any(w in vl for w in ("none","no","nil","0","nahi")):    return "No Livestock"
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Output builders
# ──────────────────────────────────────────────────────────────────────────────

def build_cat_data(counts: dict, total: int):
    """Build graphData list sorted desc by count, with pct and displayValue."""
    out = []
    for label, count in sorted(counts.items(), key=lambda x: -x[1]):
        if count <= 0 or not label:
            continue
        pct = round(count / total * 100, 1) if total else 0
        out.append({
            "indicator":    label,
            "value":        pct,
            "count":        count,
            "displayValue": f"{pct}% ({count} HH)",
        })
    return out

def build_yn_data(yn_dict: dict, total: int):
    """Build grouped-bar graphData {indicator, Yes%, No%, YesCount, NoCount}."""
    out = []
    for label, (yc, nc) in yn_dict.items():
        if yc + nc == 0:
            continue
        out.append({
            "indicator": label,
            "Yes":       round(yc / total * 100, 1) if total else 0,
            "No":        round(nc / total * 100, 1) if total else 0,
            "YesCount":  yc,
            "NoCount":   nc,
        })
    return out

def make_cat(name, chart_type, graph_data):
    if not graph_data:
        return None
    return {"category": name, "type": chart_type, "graphData": graph_data}


def parse_number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("₹", "").replace("$", "")
    text = text.replace(" ", "")
    try:
        return float(text)
    except ValueError:
        return None


def extract_my_score(ws):
    """Read the VDI sheet's My Score column and return the last numeric value."""
    rows = list(ws.iter_rows(values_only=True))
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row):
            if value is None:
                continue
            text = nrm(str(value))
            if text in {"myscore", "my score"}:
                score_col = col_idx
                values = []
                for data_row in rows[row_idx + 1:]:
                    if score_col >= len(data_row):
                        continue
                    cell = data_row[score_col]
                    if cell in (None, ""):
                        continue
                    num = parse_number(cell)
                    if num is not None:
                        values.append(num)
                return values[-1] if values else None
    return None


def extract_population_summary(gp_rows):
    targets = [
        ("Total Population", "totalPopulation"),
        ("Total Male Population", "malePopulation"),
        ("Total Female Population", "femalePopulation"),
    ]

    def find_question_key(row):
        for key in row.keys():
            kn = nrm(key)
            if kn in {"question", "questionnaire", "indicator", "parameter", "field", "attribute", "item"}:
                return key
            # also check if key contains any of these tokens (handles e.g. "Questionnaire: Gram Panchayat Level")
            if kn and any(kn.startswith(t) or t in kn for t in ("questionnaire", "question", "indicator")):
                return key
        return None

    def find_response_key(row):
        for key in row.keys():
            kn = nrm(key)
            if kn in {"response", "answer", "value", "value1", "result", "score", "count", "population", "total"}:
                return key
            if kn and (kn.startswith("response") or kn.startswith("answer") or kn == "value"):
                return key
        return None

    def find_value_for_question(rows, target_text):
        target_n = nrm(target_text)
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            question_key = find_question_key(row)
            response_key = find_response_key(row)
            question_value = None
            if question_key:
                question_value = row.get(question_key)
            else:
                # fallback when first column is the question text and second is the response
                values = [v for v in row.values() if v is not None and str(v).strip()]
                if len(values) >= 2:
                    question_value = values[0]
                    response_key = list(row.keys())[list(row.values()).index(values[1])]

            if question_value is None or response_key is None:
                continue
            if nrm(question_value) == target_n or target_n in nrm(question_value):
                response_value = row.get(response_key)
                num = parse_number(response_value)
                if num is not None:
                    return num
        return None

    values = {}
    for target_text, key in targets:
        found = find_value_for_question(gp_rows, target_text)
        if found is None:
            return None
        values[key] = found

    total_population = values.get("totalPopulation")
    male_population = values.get("malePopulation")
    female_population = values.get("femalePopulation")
    if total_population is None or total_population <= 0:
        return None
    if male_population is None or female_population is None:
        return None

    male_percentage = round((male_population / total_population) * 100, 2)
    female_percentage = round((female_population / total_population) * 100, 2)
    return {
        "totalPopulation": total_population,
        "malePopulation": male_population,
        "femalePopulation": female_population,
        "malePercentage": male_percentage,
        "femalePercentage": female_percentage,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Grouped yes/no helper for multiple columns
# ──────────────────────────────────────────────────────────────────────────────

def yn_multi(rows, total, items):
    """
    items = list of (display_label, [keyword, ...])
    Returns grouped-bar graphData.
    """
    yn = {}
    for label, kws in items:
        vals = find_col(rows, *kws)
        if not vals:
            continue
        yc = sum(1 for v in vals if is_yes(v))
        nc = sum(1 for v in vals if is_no(v))
        if yc + nc > 0:
            yn[label] = (yc, nc)
    return build_yn_data(yn, total)


# ──────────────────────────────────────────────────────────────────────────────
# Category processors
# ──────────────────────────────────────────────────────────────────────────────

def cat_population(rows, total):
    counts = Counter()
    # Try a single Gender column
    vals = find_col(rows, "gender", "sex", "gender of head", "sex of head",
                    "sex of household head", "male female", "ling")
    if vals:
        for v in vals:
            lbl = map_gender(v)
            if lbl:
                counts[lbl] += 1
    # Fallback: try separate Male Members / Female Members columns
    if not counts:
        mc = find_col(rows, "no of male", "male members", "no male", "males", "number of male")
        fc = find_col(rows, "no of female", "female members", "no female", "females", "number of female")
        if mc:
            try: counts["Male"] = int(sum(float(v) for v in mc if v is not None))
            except: pass
        if fc:
            try: counts["Female"] = int(sum(float(v) for v in fc if v is not None))
            except: pass
    return make_cat("Population", "pie", build_cat_data(counts, sum(counts.values())))


def cat_building(rows, total):
    counts = Counter()
    # House type (categorical)
    ht_vals = find_col(rows, "house type", "type of house", "construction type",
                       "housing type", "type of construction", "building type",
                       "house category", "wall material", "type of dwelling")
    if ht_vals:
        for v in ht_vals:
            lbl = map_house(v)
            if lbl:
                counts[lbl] += 1

    # Toilet – yes/no column
    t_vals = find_col(rows, "toilet", "sanitation facility", "toilet available",
                      "toilet facility", "has toilet", "toilet status")
    if t_vals:
        y = sum(1 for v in t_vals if is_yes(v))
        if y:
            counts["Toilet Available"] = y

    # Roof type if present
    r_vals = find_col(rows, "roof type", "roof material", "roof")
    if r_vals:
        for v in r_vals:
            vl = str(v).strip().lower()
            if any(w in vl for w in ("rcc","concrete","flat")):   counts["RCC Roof"] += 1
            elif any(w in vl for w in ("tin","gi sheet","metal")): counts["Tin/GI Roof"] += 1
            elif any(w in vl for w in ("thatch","grass","straw")): counts["Thatched Roof"] += 1
            elif any(w in vl for w in ("tile","khaprail")):       counts["Tile Roof"] += 1

    return make_cat("Building Details", "hbar", build_cat_data(counts, total))


def cat_socioeconomic(rows, total):
    counts = Counter()
    bpl_vals = find_col(rows, "bpl apl", "bpl", "apl", "economic category",
                        "economic status", "category", "ration card", "income category")
    if bpl_vals:
        for v in bpl_vals:
            vl = str(v).strip().lower()
            if "bpl" in vl:   counts["BPL"] += 1
            elif "apl" in vl: counts["APL"] += 1
    occ_vals = find_col(rows, "main occupation", "occupation", "occupation of head",
                        "primary occupation", "work type", "livelihood", "employment",
                        "source of income")
    if occ_vals:
        for v in occ_vals:
            lbl = map_occupation(v)
            if lbl:
                counts[lbl] += 1
    return make_cat("Socio-Economic Status", "hbar", build_cat_data(counts, total))


def cat_education(rows, total):
    counts = Counter()
    lit_vals = find_col(rows, "literacy", "literate", "read write",
                        "literacy status", "can read", "sakshar")
    if lit_vals:
        y = sum(1 for v in lit_vals if is_yes(v) or
                any(w in str(v).lower() for w in ("literate","read","sakshar","yes")))
        n = sum(1 for v in lit_vals if is_no(v) or
                any(w in str(v).lower() for w in ("illiterate","nirakshar","no")))
        if y: counts["Literate"] = y
        if n: counts["Illiterate"] = n

    edu_vals = find_col(rows, "education level", "qualification", "highest education",
                        "max education", "education qualification", "education",
                        "educational qualification")
    if edu_vals:
        for v in edu_vals:
            lbl = map_education(v)
            if lbl and lbl not in ("Literate (others)",):
                counts[lbl] += 1

    sch_vals = find_col(rows, "school going", "school attendance", "attending school",
                        "child school", "children school")
    if sch_vals:
        y = sum(1 for v in sch_vals if is_yes(v))
        if y: counts["School Going Children"] = y

    return make_cat("Education", "bar", build_cat_data(counts, total))


def cat_transport(rows, total):
    vals = find_col(rows, "mode of transport", "mode of travel", "transport",
                    "transportation", "vehicle", "commute", "travel mode", "mode")
    if not vals:
        return None
    counts = Counter()
    for v in vals:
        lbl = map_transport(v)
        if lbl:
            counts[lbl] += 1
    return make_cat("Transportation", "pie", build_cat_data(counts, sum(counts.values()) or total))


def cat_water(rows, total):
    counts = Counter()
    ws_vals = find_col(rows, "source of drinking water", "drinking water source",
                       "water source", "water supply", "source of water",
                       "drinking water", "main water source")
    if ws_vals:
        for v in ws_vals:
            lbl = map_water(v)
            if lbl:
                counts[lbl] += 1

    for label, kws in [
        ("Toilet Facility",    ["toilet","has toilet","toilet available","sanitation"]),
        ("Drainage Available", ["drainage","drain","nali"]),
        ("Waste Disposal",     ["solid waste","waste disposal","garbage","waste management"]),
    ]:
        cv = find_col(rows, *kws)
        if cv:
            y = sum(1 for v in cv if is_yes(v))
            if y:
                counts[label] = y

    return make_cat("Water & Sanitation", "hbar", build_cat_data(counts, total))


def cat_infrastructure(hh_rows, gp_rows, total):
    rows = gp_rows if gp_rows else hh_rows
    tot  = len(rows) or total
    counts = Counter()
    infra = [
        ("All-Weather Road",  ["road","all weather road","road connectivity","pucca road"]),
        ("Street Lights",     ["street light","street lamp","lighting","bijli"]),
        ("School",            ["school","primary school","education facility","vidyalay"]),
        ("Health Centre",     ["health centre","health center","hospital","chc","phc","dispensary"]),
        ("Community Hall",    ["community hall","panchayat bhawan","samudayik","sabha bhawan"]),
        ("Anganwadi",         ["anganwadi","awc","icds"]),
        ("Bank/ATM",          ["bank","atm","financial"]),
        ("Library",           ["library","pustakalay","reading room"]),
    ]
    for label, kws in infra:
        v = find_col(rows, *kws)
        if not v:
            continue
        y = sum(1 for x in v if is_yes(x))
        if y:
            counts[label] = y
    return make_cat("Infrastructure", "hbar", build_cat_data(counts, tot))


def cat_digital(rows, total):
    data = yn_multi(rows, total, [
        ("Smartphone",       ["smartphone","mobile phone","cell phone","mobile"]),
        ("Internet Access",  ["internet","net access","internet access"]),
        ("Digital Payments", ["digital payment","upi","paytm","online payment","gpay","bhim"]),
        ("Social Media",     ["social media","facebook","whatsapp","instagram"]),
        ("Smart Village",    ["smart village","desire smart","smart vill"]),
    ])
    return make_cat("Digital Connectivity", "grouped-bar", data)


def cat_electricity(rows, total):
    counts = Counter()
    for label, kws in [
        ("Electricity Connection", ["electricity connection","electrical connection","electrified","has electricity","power connection"]),
        ("Meter Installed",        ["meter","electricity meter","meter installed","metered"]),
    ]:
        v = find_col(rows, *kws)
        if v:
            y = sum(1 for x in v if is_yes(x))
            if y:
                counts[label] = y
    # Supply hours – group into ranges
    h_vals = find_col(rows, "electricity hours","power hours","supply hours","hours of supply","hours of electricity")
    if h_vals:
        hr_counts = Counter()
        for v in h_vals:
            try:
                h = float(str(v).strip())
                if h >= 20:    hr_counts["≥20 hrs/day"] += 1
                elif h >= 12:  hr_counts["12-20 hrs/day"] += 1
                elif h >= 6:   hr_counts["6-12 hrs/day"] += 1
                else:          hr_counts["<6 hrs/day"] += 1
            except (ValueError, TypeError):
                pass
        counts.update(hr_counts)
    return make_cat("Electricity", "hbar", build_cat_data(counts, total))


def cat_fuel(rows, total):
    vals = find_col(rows, "cooking fuel", "fuel type", "fuel", "cooking energy",
                    "type of fuel", "fuel used", "energy for cooking")
    if not vals:
        return None
    counts = Counter()
    for v in vals:
        lbl = map_fuel(v)
        if lbl:
            counts[lbl] += 1
    return make_cat("Clean Fuel", "pie", build_cat_data(counts, sum(counts.values()) or total))


def cat_health(rows, total):
    data = yn_multi(rows, total, [
        ("Health Insurance",       ["health insurance","insurance","pmjay","ayushman","pm jan arogya"]),
        ("Institutional Delivery", ["institutional delivery","hospital delivery","delivery","prasav"]),
        ("Child Vaccination",      ["vaccination","immunisation","immunization","tikaakaran"]),
        ("Anganwadi Access",       ["anganwadi","awc","icds"]),
        ("Toilet Usage",           ["toilet usage","open defecation free","odf","toilet use"]),
    ])
    return make_cat("Health & Nutrition", "grouped-bar", data)


def cat_agriculture(rows, total):
    counts = Counter()

    def normalize_income_source(value):
        if value is None:
            return None
        text = str(value).strip().lower()
        if not text:
            return None

        parts = []
        for part in re.split(r"\s*\+\s*", text):
            cleaned = re.split(r"\s*[\/&,]\s*", part.strip())
            parts.extend([p for p in cleaned if p.strip()])

        for part in parts:
            p = re.sub(r"[^a-z]+", "", part.lower())
            if not p:
                continue
            if p in {"agri", "agriculture", "farming", "farm", "cultivation"}:
                return "Agriculture"
            if p in {"govtjob", "govt", "governmentjob", "government", "sarkari", "service", "job", "govtjobservice"}:
                return "Government Job"
            if p in {"pvt", "private", "privatejob", "pvtjob"}:
                return "Private Job"
            if p in {"labour", "labor", "dailywage", "manual", "mazdoor"}:
                return "Labour"
            if p in {"business", "selfemployed", "selfempd", "selfemployment", "shop", "trade", "entrepreneur", "businessselfempd"}:
                return "Business / Self Employed"

        return "Other"

    income_vals = find_col(rows, "main source of income", "source of income", "main occupation",
                           "occupation of head", "primary occupation", "livelihood",
                           "income source", "occupation")
    if income_vals:
        for v in income_vals:
            if v is None:
                continue
            raw_parts = [p.strip() for p in str(v).split("+") if p.strip()]
            for part in raw_parts:
                label = normalize_income_source(part)
                if label:
                    counts[label] += 1

    if not counts:
        return None

    return make_cat("Household Source of Income", "hbar", build_cat_data(counts, total))


def cat_livestock(rows, total):
    # Try a single column that names the animal
    main_vals = find_col(rows, "livestock owned","livestock","animal husbandry",
                         "cattle owned","animals")
    if main_vals:
        counts = Counter()
        for v in main_vals:
            lbl = map_livestock(v)
            if lbl:
                counts[lbl] += 1
        if counts:
            return make_cat("Livestock / Animal Husbandry", "pie",
                            build_cat_data(counts, sum(counts.values())))

    # Fallback: individual animal yes/no columns
    animals = [
        ("Cow",      ["cow","gaay","gai"]),
        ("Buffalo",  ["buffalo","bhains","bull","ox"]),
        ("Goat",     ["goat","bakri","sheep"]),
        ("Poultry",  ["poultry","hen","chicken","murgi","duck"]),
        ("Pig",      ["pig","swine","suar"]),
    ]
    counts = Counter()
    any_found = False
    for label, kws in animals:
        av = find_col(rows, *kws)
        if not av:
            continue
        any_found = True
        y = sum(1 for v in av if is_yes(v) or (
            str(v).strip().isdigit() and int(str(v).strip()) > 0))
        if y:
            counts[label] = y
    if not any_found:
        return None
    no_lv = total - sum(counts.values())
    if no_lv > 0:
        counts["No Livestock"] = no_lv
    return make_cat("Livestock / Animal Husbandry", "pie",
                    build_cat_data(counts, total))


def cat_social(rows, total):
    data = yn_multi(rows, total, [
        ("SHG Membership",        ["shg","self help group","shg member","mahila mandal"]),
        ("Govt Scheme Awareness",  ["government scheme","scheme awareness","aware of scheme","yojana"]),
        ("Skill Training",         ["skill training","vocational","skill development","pratikshan"]),
        ("Digital Awareness",      ["digital awareness","digital literacy","digital knowledge"]),
        ("Gram Sabha Attendance",  ["gram sabha","gramSabha","sabha","community meeting"]),
    ])
    return make_cat("Socially Inclusive & Awareness", "grouped-bar", data)


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

def process(filename):
    wb = load_workbook(filename, data_only=True)

    # Load Household Survey sheet
    hh_rows = []
    for name in ("HOUSEHOLD SURVEY", "HOUSEHOLD", "HH SURVEY", "HH", "SURVEY",
                 "HOUSEHOLD DATA", "HH DATA"):
        try:
            sheet = worksheet(wb, name)
            r = transposed_households(sheet) or table_rows(sheet)
            if len(r) >= 5:
                hh_rows = r
                break
        except ValueError:
            continue

    # Fallback: biggest sheet with ≥5 columns and ≥10 rows (that isn't named VDI/GP)
    if not hh_rows:
        best = None
        for sname in wb.sheetnames:
            if any(k in nrm(sname) for k in ("vdi", "gram", "panchayat", "gp")):
                continue
            r = table_rows(wb[sname])
            if r and len(r) >= 10 and len(list(r[0].keys())) >= 5:
                if best is None or len(r) > len(best):
                    best = r
        hh_rows = best or []

    # Load Gram Panchayat sheet
    gp_rows = []
    for name in ("GRAM PANCHAYAT", "GP", "GRAM PANCHAYAT DATA", "VILLAGE DATA",
                 "PANCHAYAT", "GRAMPANCHAYAT"):
        try:
            r = table_rows(worksheet(wb, name))
            if r:
                gp_rows = r
                break
        except ValueError:
            continue

    total = len(hh_rows)
    population_summary = extract_population_summary(gp_rows)
    my_score = None
    try:
        vdi_sheet = worksheet(wb, "VDI")
        my_score = extract_my_score(vdi_sheet)
    except Exception:
        pass

    cats = []
    for fn in [
        lambda: cat_population(hh_rows, total),
        lambda: cat_building(hh_rows, total),
        lambda: cat_socioeconomic(hh_rows, total),
        lambda: cat_education(hh_rows, total),
        lambda: cat_transport(hh_rows, total),
        lambda: cat_water(hh_rows, total),
        lambda: cat_infrastructure(hh_rows, gp_rows, total),
        lambda: cat_digital(hh_rows, total),
        lambda: cat_electricity(hh_rows, total),
        lambda: cat_fuel(hh_rows, total),
        lambda: cat_health(hh_rows, total),
        lambda: cat_agriculture(hh_rows, total),
        lambda: cat_livestock(hh_rows, total),
        lambda: cat_social(hh_rows, total),
    ]:
        try:
            c = fn()
            if c:
                cats.append(c)
        except Exception:
            pass  # never crash on a single category

    result = {"categories": cats}
    if my_score is not None:
        result["summary"] = {"myScore": my_score}
    if population_summary:
        result["population"] = population_summary
        result["populationSummary"] = population_summary
    else:
        print("Population summary missing or invalid in Gram Panchayat sheet; expected Total Population, Total Male Population, Total Female Population.", file=sys.stderr)
    return result


if __name__ == "__main__":
    try:
        print(json.dumps(process(sys.argv[1]), default=str, ensure_ascii=False))
    except Exception as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)