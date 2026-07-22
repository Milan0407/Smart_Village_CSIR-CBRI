# -*- coding: utf-8 -*-
"""Deep-inspect key question rows to see actual household responses"""
import sys
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(path, data_only=True)
ws = wb["HOUSEHOLD SURVEY"]

# Resolve merged cells
merged = {}
for rng in ws.merged_cells.ranges:
    top = ws.cell(rng.min_row, rng.min_col).value
    for r in range(rng.min_row, rng.max_row + 1):
        for c in range(rng.min_col, rng.max_col + 1):
            merged[(r, c)] = top

rows = []
for ri, row in enumerate(ws.iter_rows(values_only=True)):
    resolved = [merged.get((ri+1, ci+1), v) for ci, v in enumerate(row)]
    rows.append(resolved)

# Row 4 = HOUSEHOLD 1, HOUSEHOLD 2 etc (col headers)
# Row 5 = survey index (1.0, 2.0...)
# Row 8 = DIMENSIONS, INDICATORS header
# Data cols start at index 2 (col C onwards per household)

print("=== Row 4 (household headers): first 10 cols ===")
print(rows[3][:12])

print("\n=== Row 8 (column headers): first 10 cols ===")
print(rows[7][:12])

print("\n=== Row 9 (first data row - col 0 is S.NO, col 1 is DIMENSIONS, col 2 is INDICATORS): ===")
print(rows[8][:12])

# Find how many household data columns exist
# Row 4 has "HOUSEHOLD 1", "HOUSEHOLD 2" etc
hh_labels = [str(v).strip() for v in rows[3] if v and "HOUSEHOLD" in str(v).upper()]
print(f"\nNumber of households in row 4: {len(hh_labels)}")
print(f"First 5: {hh_labels[:5]}")

# Key rows to inspect (0-indexed: row 13=female, row 14=male, row 25=house type, etc)
key_rows = [
    (12, "Number of Females"),
    (13, "Number of Males"),
    (24, "Type of House"),
    (25, "Area of House"),
    (26, "Wall Type"),
    (27, "Roof Type"),
    (32, "Main Source of Income"),
    (38, "No of Literate"),
    (44, "Public Transport"),
    (52, "Smart Phone"),
    (62, "Electricity"),
    (66, "Clean Fuel"),
    (89, "Livestock"),
]

print("\n=== KEY ROWS - First 10 household values ===")
for ridx, label in key_rows:
    if ridx < len(rows):
        row = rows[ridx]
        # col 0 = S.NO, col 1 = DIMENSIONS, col 2 = INDICATORS (full question), cols 3+ = household data
        indicator = str(row[2]).strip()[:50] if len(row) > 2 else ""
        vals = [str(v).strip() if v is not None else "-" for v in row[3:13]]
        print(f"\n  Row {ridx+1}: [{label}] | Indicator: {indicator}")
        print(f"    HH values: {vals}")
